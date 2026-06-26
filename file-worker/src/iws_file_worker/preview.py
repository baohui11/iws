from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import requests
import xlrd

from .config import Settings
from .models import FileRow
from .storage import ObjectStorage

OFFICE_TO_PDF_EXT = {"doc", "docx", "ppt", "pptx"}
EXCEL_EXT = {"xlsx", "xls"}
CSV_EXT = {"csv", "tsv"}
DIRECT_PREVIEW_EXT = {
    "pdf",
    "png",
    "jpg",
    "jpeg",
    "gif",
    "webp",
    "bmp",
    "svg",
    "txt",
    "log",
    "md",
    "markdown",
    "mp3",
    "wav",
    "aac",
    "ogg",
    "m4a",
    "flac",
    "mp4",
    "webm",
    "mov",
}

MAX_SHEETS = 3
MAX_ROWS = 200
MAX_COLS = 50
MAX_CELL_CHARS = 500


@dataclass(frozen=True)
class PreviewResult:
    status: str
    storage_key: str | None = None
    output: dict[str, Any] | None = None
    error_code: str | None = None
    error_message: str | None = None


def normalize_ext(value: str | None) -> str:
    return (value or "").strip().lower().lstrip(".")


def preview_artifact_key(file: FileRow, suffix: str) -> str:
    return f"{file.project_id}/preview/{file.id}{suffix}"


def cell_value(value: Any) -> str | int | float | bool | None:
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    if len(text) > MAX_CELL_CHARS:
        return text[:MAX_CELL_CHARS] + "..."
    return text


def build_xlsx_preview(data: bytes) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(
        io.BytesIO(data),
        read_only=True,
        data_only=True,
    )
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.worksheets[:MAX_SHEETS]:
        rows: list[list[Any]] = []
        for row in sheet.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True):
            rows.append([cell_value(v) for v in row])
        sheets.append(
            {
                "name": sheet.title,
                "rows": rows,
                "truncated": sheet.max_row > MAX_ROWS or sheet.max_column > MAX_COLS,
            }
        )
    return {"version": 1, "kind": "spreadsheet", "sheets": sheets}


def build_xls_preview(data: bytes) -> dict[str, Any]:
    workbook = xlrd.open_workbook(file_contents=data)
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.sheets()[:MAX_SHEETS]:
        rows: list[list[Any]] = []
        row_count = min(sheet.nrows, MAX_ROWS)
        col_count = min(sheet.ncols, MAX_COLS)
        for r in range(row_count):
            rows.append([cell_value(sheet.cell_value(r, c)) for c in range(col_count)])
        sheets.append(
            {
                "name": sheet.name,
                "rows": rows,
                "truncated": sheet.nrows > MAX_ROWS or sheet.ncols > MAX_COLS,
            }
        )
    return {"version": 1, "kind": "spreadsheet", "sheets": sheets}


def decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def build_csv_preview(data: bytes, delimiter: str) -> dict[str, Any]:
    text = decode_text(data)
    rows: list[list[str]] = []
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    truncated = False
    for row_idx, row in enumerate(reader):
        if row_idx >= MAX_ROWS:
            truncated = True
            break
        if len(row) > MAX_COLS:
            truncated = True
        rows.append([str(cell_value(v) or "") for v in row[:MAX_COLS]])
    return {
        "version": 1,
        "kind": "spreadsheet",
        "sheets": [{"name": "Sheet1", "rows": rows, "truncated": truncated}],
    }


def convert_office_to_pdf(
    settings: Settings,
    *,
    file_name: str,
    data: bytes,
) -> bytes:
    url = settings.gotenberg_url.rstrip("/") + "/forms/libreoffice/convert"
    response = requests.post(
        url,
        files={"files": (file_name, data)},
        timeout=120,
    )
    response.raise_for_status()
    return response.content


class PreviewProcessor:
    def __init__(self, settings: Settings, storage: ObjectStorage) -> None:
        self.settings = settings
        self.storage = storage

    def process(self, file: FileRow) -> PreviewResult:
        ext = normalize_ext(file.file_ext or Path(file.file_name).suffix)
        if ext in OFFICE_TO_PDF_EXT:
            source = self.storage.get_bytes(file.source_storage_key)
            pdf = convert_office_to_pdf(
                self.settings,
                file_name=file.file_name,
                data=source.body,
            )
            key = preview_artifact_key(file, ".pdf")
            self.storage.put_bytes(key, pdf, "application/pdf")
            return PreviewResult(
                status="ready",
                storage_key=key,
                output={"version": 1, "kind": "pdf", "bytes": len(pdf)},
            )

        if ext in EXCEL_EXT:
            source = self.storage.get_bytes(file.source_storage_key)
            preview = (
                build_xlsx_preview(source.body)
                if ext == "xlsx"
                else build_xls_preview(source.body)
            )
            key = preview_artifact_key(file, ".preview.json")
            body = json.dumps(preview, ensure_ascii=False).encode("utf-8")
            self.storage.put_bytes(key, body, "application/json; charset=utf-8")
            return PreviewResult(status="ready", storage_key=key, output=preview)

        if ext in CSV_EXT:
            source = self.storage.get_bytes(file.source_storage_key)
            preview = build_csv_preview(source.body, "\t" if ext == "tsv" else ",")
            key = preview_artifact_key(file, ".preview.json")
            body = json.dumps(preview, ensure_ascii=False).encode("utf-8")
            self.storage.put_bytes(key, body, "application/json; charset=utf-8")
            return PreviewResult(status="ready", storage_key=key, output=preview)

        if ext in DIRECT_PREVIEW_EXT:
            return PreviewResult(
                status="skipped",
                output={
                    "version": 1,
                    "reason": "source_file_is_directly_previewable",
                },
            )

        return PreviewResult(
            status="skipped",
            output={"version": 1, "reason": "unsupported_preview_type"},
        )
